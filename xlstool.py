"""
Excel workbook utilities: sheet summaries and related helpers.
Supports legacy .xls (xlrd), .xlsx / .xlsm (openpyxl), and comma-separated ``.csv`` for ``--sum``. Run: python xlstool.py -h

``--dup`` / ``-dup`` finds duplicate rows in a CSV using a composite key built from
``--keys``: either 1-based column indexes with optional ranges (``1,3,5-7``), or
comma-separated header names (first row is the header; data starts on row 2).

``--merge-csv DIR`` concatenates files under ``DIR`` matching ``--merge-pattern`` (default
``*.csv``), sorted by name: the first file supplies the header row; each later file drops
its first row and appends the rest. Default output: ``<parent>/<DIRname>_merged.csv``
(sibling of ``DIR``); use ``-o`` for another path. Output is always UTF-8 with BOM.

Sheet row/column counts are the tight bounding box of cells with non-null values;
strings that are empty or whitespace-only are ignored (numeric zero counts).

-xc / --xc exports sheet(s) to CSV files named ``{workbook_stem}_{sheetName}.csv``
next to the workbook (sheet names are sanitized for the filesystem). Each CSV is the
minimal non-empty cell rectangle (same rule as ``--sum``), not the full allocated grid.
``--sheets`` / ``-sh`` accepts sheet names and/or 1-based indexes (same order as ``--sum``); use ``#2`` or ``@2`` to pick by index when a sheet is named like a number.
With ``--sum``, ``-sh`` on workbooks limits which sheets are listed and adds **column names** (top row of the tight non-empty rectangle; empty header cells show as ``(A)``, ``(B)``, and so on). For a ``.csv``, ``--sum`` always includes those column names (``-sh`` only filters whether that sheet is listed).
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import math
from collections import defaultdict
from pathlib import Path
from typing import Iterable, NamedTuple


class SheetStat(NamedTuple):
    index: int
    name: str
    rows: int
    columns: int
    column_labels: tuple[str, ...] | None


def _normalize_argv(argv: list[str]) -> list[str]:
    """Map literal '-sum' / '-xc' / '-sh' to long forms (argparse short-option clusters)."""
    out: list[str] = []
    for a in argv:
        if a == "-sum":
            out.append("--sum")
        elif a == "-xc":
            out.append("--xc")
        elif a == "-sh":
            out.append("--sheets")
        elif a == "-dup":
            out.append("--dup")
        elif a == "-merge-csv":
            out.append("--merge-csv")
        else:
            out.append(a)
    return out


_WIN_FILE_INVALID = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def _sanitize_sheet_for_filename(sheet_name: str, max_component: int = 120) -> str:
    """Make a single path component safe on Windows and strip trailing dots/spaces."""
    s = _WIN_FILE_INVALID.sub("_", sheet_name)
    s = s.rstrip(" .")
    if not s:
        s = "sheet"
    if len(s) > max_component:
        s = s[:max_component].rstrip(" .")
        if not s:
            s = "sheet"
    return s


def _unique_csv_path(out_dir: Path, workbook_stem: str, sheet_name: str, used_lower: set[str]) -> Path:
    safe = _sanitize_sheet_for_filename(sheet_name)
    base = f"{workbook_stem}_{safe}"
    candidate = f"{base}.csv"
    n = 2
    while candidate.lower() in used_lower:
        candidate = f"{base}_{n}.csv"
        n += 1
    used_lower.add(candidate.lower())
    return out_dir / candidate


def _cell_has_content(val: object) -> bool:
    """True if the cell value counts as non-null and non-empty (whitespace-only strings are empty)."""
    if val is None:
        return False
    if isinstance(val, str):
        return bool(val.strip())
    if isinstance(val, float) and math.isnan(val):
        return False
    return True


def _xlrd_sheet_nonempty_extent(sh) -> tuple[int, int, int, int] | None:
    """Inclusive 0-based (min_r, max_r, min_c, max_c) over non-empty cells, or None if none."""
    import xlrd

    min_r = min_c = None
    max_r = max_c = None
    for r in range(sh.nrows):
        for c in range(sh.ncols):
            typ = sh.cell_type(r, c)
            if typ in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                continue
            v = sh.cell_value(r, c)
            if not _cell_has_content(v):
                continue
            if min_r is None:
                min_r = max_r = r
                min_c = max_c = c
            else:
                if r < min_r:
                    min_r = r
                if r > max_r:
                    max_r = r
                if c < min_c:
                    min_c = c
                if c > max_c:
                    max_c = c
    if min_r is None:
        return None
    return min_r, max_r, min_c, max_c


def _xlrd_sheet_nonempty_bounds(sh) -> tuple[int, int]:
    """Row/column counts of the minimal non-empty rectangle; (0, 0) if none."""
    ext = _xlrd_sheet_nonempty_extent(sh)
    if ext is None:
        return 0, 0
    min_r, max_r, min_c, max_c = ext
    return max_r - min_r + 1, max_c - min_c + 1


def _label_for_header(v: object, excel_col_1based: int) -> str:
    """Display string for a header cell; empty values use Excel column letter in parentheses."""
    from openpyxl.utils import get_column_letter

    letter = get_column_letter(excel_col_1based)
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return f"({letter})"
    if isinstance(v, str):
        s = v.strip()
        return s if s else f"({letter})"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    s = str(v).strip()
    return s if s else f"({letter})"


def _xlrd_header_labels(book, sh, ext: tuple[int, int, int, int]) -> tuple[str, ...]:
    import xlrd

    min_r, _max_r, min_c, max_c = ext
    out: list[str] = []
    for c in range(min_c, max_c + 1):
        typ = sh.cell_type(min_r, c)
        if typ in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            v: object = None
        else:
            v = sh.cell_value(min_r, c)
            if typ == xlrd.XL_CELL_TEXT and isinstance(v, str) and not v.strip():
                v = None
            elif typ == xlrd.XL_CELL_NUMBER and isinstance(v, float) and math.isnan(v):
                v = None
        out.append(_label_for_header(v, c + 1))
    return tuple(out)


def _openpyxl_header_labels(ws, ext: tuple[int, int, int, int]) -> tuple[str, ...]:
    min_r, _max_r, min_c, max_c = ext
    row = next(
        ws.iter_rows(
            min_row=min_r,
            max_row=min_r,
            min_col=min_c,
            max_col=max_c,
            values_only=True,
        )
    )
    return tuple(_label_for_header(v, min_c + i) for i, v in enumerate(row))


def _summarize_xls(path: Path, sheet_tokens: list[str] | None) -> tuple[list[SheetStat], list[str]]:
    import xlrd

    bad: list[str] = []
    book = xlrd.open_workbook(str(path), on_demand=True)
    try:
        names = [book.sheet_names()[i] for i in range(book.nsheets)]
        include: set[str] | None = None
        detailed = sheet_tokens is not None
        if sheet_tokens is not None:
            resolved, bad = _resolve_sheet_tokens(sheet_tokens, names)
            include = resolved
        stats: list[SheetStat] = []
        for i in range(book.nsheets):
            name = names[i]
            if include is not None and name not in include:
                continue
            sh = book.sheet_by_index(i)
            ext = _xlrd_sheet_nonempty_extent(sh)
            if ext is None:
                rows, cols = 0, 0
            else:
                min_r, max_r, min_c, max_c = ext
                rows, cols = max_r - min_r + 1, max_c - min_c + 1
            labels: tuple[str, ...] | None
            if detailed:
                labels = () if ext is None else _xlrd_header_labels(book, sh, ext)
            else:
                labels = None
            stats.append(SheetStat(index=i + 1, name=name, rows=rows, columns=cols, column_labels=labels))
        return stats, bad
    finally:
        book.release_resources()


def _openpyxl_sheet_nonempty_extent(ws) -> tuple[int, int, int, int] | None:
    """Inclusive 1-based (min_row, max_row, min_col, max_col) over non-empty cells, or None if none."""
    min_r = min_c = None
    max_r = max_c = None
    for row in ws.iter_rows():
        for cell in row:
            if not _cell_has_content(cell.value):
                continue
            r, c = cell.row, cell.column
            if min_r is None:
                min_r = max_r = r
                min_c = max_c = c
            else:
                if r < min_r:
                    min_r = r
                if r > max_r:
                    max_r = r
                if c < min_c:
                    min_c = c
                if c > max_c:
                    max_c = c
    if min_r is None:
        return None
    return min_r, max_r, min_c, max_c


def _openpyxl_sheet_nonempty_bounds(ws) -> tuple[int, int]:
    """Row/column counts of the minimal non-empty rectangle; (0, 0) if none."""
    ext = _openpyxl_sheet_nonempty_extent(ws)
    if ext is None:
        return 0, 0
    min_r, max_r, min_c, max_c = ext
    return max_r - min_r + 1, max_c - min_c + 1


def _summarize_xlsx(path: Path, sheet_tokens: list[str] | None) -> tuple[list[SheetStat], list[str]]:
    import openpyxl

    bad: list[str] = []
    detailed = sheet_tokens is not None
    wb = openpyxl.load_workbook(path, read_only=not detailed, data_only=True)
    try:
        names = list(wb.sheetnames)
        include: set[str] | None = None
        if sheet_tokens is not None:
            resolved, bad = _resolve_sheet_tokens(sheet_tokens, names)
            include = resolved
        stats: list[SheetStat] = []
        for i, name in enumerate(names, start=1):
            if include is not None and name not in include:
                continue
            ws = wb[name]
            ext = _openpyxl_sheet_nonempty_extent(ws)
            if ext is None:
                rows, cols = 0, 0
            else:
                min_r, max_r, min_c, max_c = ext
                rows, cols = max_r - min_r + 1, max_c - min_c + 1
            labels: tuple[str, ...] | None
            if detailed:
                labels = () if ext is None else _openpyxl_header_labels(ws, ext)
            else:
                labels = None
            stats.append(SheetStat(index=i, name=name, rows=rows, columns=cols, column_labels=labels))
        return stats, bad
    finally:
        wb.close()


def summarize_workbook(path: Path, sheet_tokens: list[str] | None = None) -> tuple[list[SheetStat], list[str]]:
    suf = path.suffix.lower()
    if suf == ".xls":
        return _summarize_xls(path, sheet_tokens)
    if suf in (".xlsx", ".xlsm"):
        return _summarize_xlsx(path, sheet_tokens)
    if suf == ".csv":
        return _summarize_csv(path, sheet_tokens)
    raise ValueError(
        f"Unsupported file type {path.suffix!r} for --sum (use .csv, .xls, .xlsx, or .xlsm)"
    )


def _parse_sheet_tokens(spec: str | None) -> list[str] | None:
    if spec is None or not str(spec).strip():
        return None
    parts = [p.strip() for p in str(spec).split(",") if p.strip()]
    return parts or None


_INDEX_FORCED = re.compile(r"^[@#](\d+)$")


def _resolve_sheet_tokens(tokens: list[str], ordered_names: list[str]) -> tuple[set[str], list[str]]:
    """
    Map --sheets / -sh tokens to sheet names. 1-based indexes match the # column from --sum.
    Returns (resolved sheet names, list of tokens that could not be resolved).
    """
    n = len(ordered_names)
    name_set = set(ordered_names)
    resolved: set[str] = set()
    bad: list[str] = []
    for raw in tokens:
        tok = raw.strip()
        m = _INDEX_FORCED.fullmatch(tok)
        if m:
            idx = int(m.group(1))
            if 1 <= idx <= n:
                resolved.add(ordered_names[idx - 1])
            else:
                bad.append(raw)
            continue
        if tok in name_set:
            resolved.add(tok)
            continue
        if re.fullmatch(r"\d+", tok):
            idx = int(tok)
            if 1 <= idx <= n:
                resolved.add(ordered_names[idx - 1])
            else:
                bad.append(raw)
            continue
        bad.append(raw)
    return resolved, bad


def _xlrd_cell_value_for_csv(book, sh, r: int, c: int):
    import xlrd

    typ = sh.cell_type(r, c)
    v = sh.cell_value(r, c)
    if typ in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    if typ == xlrd.XL_CELL_ERROR:
        try:
            return xlrd.error_text_from_code(v)
        except Exception:
            return ""
    if typ == xlrd.XL_CELL_BOOLEAN:
        return bool(v)
    if typ == xlrd.XL_CELL_DATE:
        try:
            import xlrd.xldate

            return xlrd.xldate.xldate_as_datetime(v, book.datemode)
        except Exception:
            return v
    if typ == xlrd.XL_CELL_NUMBER:
        if isinstance(v, float) and math.isnan(v):
            return ""
        if isinstance(v, float) and v == int(v):
            return int(v)
        return v
    return v


def _csv_normalize_cell(v: object) -> object:
    if v is None:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    return v


def _write_csv_rows(path: Path, rows: Iterable[Iterable[object]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, lineterminator="\n")
        for row in rows:
            writer.writerow([_csv_normalize_cell(v) for v in row])


def _export_xls_csv(path: Path, sheet_tokens: list[str] | None, out_dir: Path) -> list[Path]:
    import xlrd

    written: list[Path] = []
    used_names: set[str] = set()
    stem = path.stem
    book = xlrd.open_workbook(str(path), on_demand=True)
    try:
        names = [book.sheet_names()[i] for i in range(book.nsheets)]
        include: set[str] | None
        if sheet_tokens is None:
            include = None
        else:
            resolved, bad = _resolve_sheet_tokens(sheet_tokens, names)
            if bad:
                print(
                    f"Warning: unknown or invalid sheet selector(s), skipped: {', '.join(bad)}",
                    file=sys.stderr,
                )
            include = resolved
        for i in range(book.nsheets):
            sh = book.sheet_by_index(i)
            name = sh.name
            if include is not None and name not in include:
                continue
            out_path = _unique_csv_path(out_dir, stem, name, used_names)
            ext = _xlrd_sheet_nonempty_extent(sh)
            if ext is None:
                _write_csv_rows(out_path, [])
            else:
                min_r, max_r, min_c, max_c = ext
                rows = (
                    [_xlrd_cell_value_for_csv(book, sh, r, c) for c in range(min_c, max_c + 1)]
                    for r in range(min_r, max_r + 1)
                )
                _write_csv_rows(out_path, rows)
            written.append(out_path)
        return written
    finally:
        book.release_resources()


def _export_xlsx_csv(path: Path, sheet_tokens: list[str] | None, out_dir: Path) -> list[Path]:
    import openpyxl

    written: list[Path] = []
    used_names: set[str] = set()
    stem = path.stem
    # read_only=False: need two passes (extent scan + bounded iter_rows) per sheet.
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        names = list(wb.sheetnames)
        include: set[str] | None
        if sheet_tokens is None:
            include = None
        else:
            resolved, bad = _resolve_sheet_tokens(sheet_tokens, names)
            if bad:
                print(
                    f"Warning: unknown or invalid sheet selector(s), skipped: {', '.join(bad)}",
                    file=sys.stderr,
                )
            include = resolved
        for name in names:
            if include is not None and name not in include:
                continue
            out_path = _unique_csv_path(out_dir, stem, name, used_names)
            ws = wb[name]
            ext = _openpyxl_sheet_nonempty_extent(ws)
            if ext is None:
                _write_csv_rows(out_path, [])
            else:
                min_r, max_r, min_c, max_c = ext
                rows = (
                    list(row)
                    for row in ws.iter_rows(
                        min_row=min_r,
                        max_row=max_r,
                        min_col=min_c,
                        max_col=max_c,
                        values_only=True,
                    )
                )
                _write_csv_rows(out_path, rows)
            written.append(out_path)
        return written
    finally:
        wb.close()


def export_workbook_csv(path: Path, sheet_tokens: list[str] | None, out_dir: Path) -> list[Path]:
    suf = path.suffix.lower()
    if suf == ".xls":
        return _export_xls_csv(path, sheet_tokens, out_dir)
    if suf in (".xlsx", ".xlsm"):
        return _export_xlsx_csv(path, sheet_tokens, out_dir)
    raise ValueError(f"Unsupported file type {path.suffix!r} for --xc (use .xls, .xlsx, or .xlsm)")


_NUMERIC_KEYS_ONLY = re.compile(r"^\s*\d+(?:\s*-\s*\d+)?(?:\s*,\s*\d+(?:\s*-\s*\d+)?)*\s*$")


def _keys_spec_is_numeric(spec: str) -> bool:
    """True if spec is only comma-separated integers and hyphen ranges (1-based column indexes)."""
    return bool(_NUMERIC_KEYS_ONLY.match(spec))


def _max_row_width(rows: list[list[str]]) -> int:
    return max((len(r) for r in rows), default=0)


def _parse_numeric_key_columns(spec: str, num_cols: int) -> list[int]:
    """1-based column indexes and inclusive ranges; returns sorted unique 0-based indices."""
    indices: set[int] = set()
    for raw in spec.split(","):
        part = raw.strip()
        if not part:
            raise ValueError("Empty segment in --keys (check commas)")
        if "-" in part:
            a, b = part.split("-", 1)
            lo = int(a.strip())
            hi = int(b.strip())
            if lo > hi:
                lo, hi = hi, lo
            for k in range(lo, hi + 1):
                idx = k - 1
                if idx < 0 or idx >= num_cols:
                    raise ValueError(f"Column {k} out of range (widest row has {num_cols} column(s))")
                indices.add(idx)
        else:
            k = int(part)
            idx = k - 1
            if idx < 0 or idx >= num_cols:
                raise ValueError(f"Column {k} out of range (widest row has {num_cols} column(s))")
            indices.add(idx)
    return sorted(indices)


def _parse_name_key_columns(spec: str, header: list[str]) -> list[int]:
    """Map comma-separated header names to 0-based column indexes (exact match, then case-insensitive)."""
    want = [p.strip() for p in spec.split(",") if p.strip()]
    if not want:
        raise ValueError("No column names in --keys")
    hdr = [h.strip() for h in header]
    cf_to_idx: dict[str, list[int]] = defaultdict(list)
    for i, h in enumerate(hdr):
        cf_to_idx[h.casefold()].append(i)
    out: list[int] = []
    for name in want:
        found: int | None = None
        for i, h in enumerate(hdr):
            if h == name:
                found = i
                break
        if found is not None:
            out.append(found)
            continue
        cands = cf_to_idx.get(name.casefold(), [])
        if len(cands) == 1:
            out.append(cands[0])
        elif len(cands) > 1:
            raise ValueError(f"Ambiguous column name {name!r}: matches multiple columns")
        raise ValueError(f"Unknown column name {name!r} (headers: {hdr})")
    return out


def _read_csv_rows(path: Path, encoding: str) -> list[list[str]]:
    with path.open(newline="", encoding=encoding) as f:
        return list(csv.reader(f))


def _csv_grid_nonempty_extent(rows: list[list[str]]) -> tuple[int, int, int, int] | None:
    """Inclusive 0-based (min_r, max_r, min_c, max_c) over cells with content (same emptiness rule as --sum)."""
    if not rows:
        return None
    width = _max_row_width(rows)
    min_r = min_c = None
    max_r = max_c = None
    for r, row in enumerate(rows):
        for c in range(width):
            val = row[c] if c < len(row) else ""
            if not _cell_has_content(val):
                continue
            if min_r is None:
                min_r = max_r = r
                min_c = max_c = c
            else:
                if r < min_r:
                    min_r = r
                if r > max_r:
                    max_r = r
                if c < min_c:
                    min_c = c
                if c > max_c:
                    max_c = c
    if min_r is None:
        return None
    return min_r, max_r, min_c, max_c


def _csv_header_labels_from_extent(rows: list[list[str]], ext: tuple[int, int, int, int]) -> tuple[str, ...]:
    min_r, _max_r, min_c, max_c = ext
    row = rows[min_r]
    return tuple(
        _label_for_header(row[c] if c < len(row) else None, c + 1) for c in range(min_c, max_c + 1)
    )


def _summarize_csv(path: Path, sheet_tokens: list[str] | None) -> tuple[list[SheetStat], list[str]]:
    """Treat a CSV as a single sheet named like the file stem (for -sh matching).

    Column names are always taken from the top row of the tight non-empty rectangle
    (same as ``--sum -sh`` on workbooks), so plain ``--sum`` on a .csv includes headers.
    """
    stem = path.stem or "CSV"
    names = [stem]
    bad: list[str] = []
    include: set[str] | None = None
    if sheet_tokens is not None:
        resolved, bad = _resolve_sheet_tokens(sheet_tokens, names)
        include = resolved
    try:
        data_rows = _read_csv_rows(path, "utf-8-sig")
    except UnicodeDecodeError as e:
        raise ValueError(f"Could not decode CSV as UTF-8 (utf-8-sig): {e}") from e
    if include is not None and stem not in include:
        return [], bad
    ext = _csv_grid_nonempty_extent(data_rows)
    if ext is None:
        rows_n, cols_n = 0, 0
    else:
        min_r, max_r, min_c, max_c = ext
        rows_n, cols_n = max_r - min_r + 1, max_c - min_c + 1
    labels: tuple[str, ...] = () if ext is None else _csv_header_labels_from_extent(data_rows, ext)
    stats = [SheetStat(index=1, name=stem, rows=rows_n, columns=cols_n, column_labels=labels)]
    return stats, bad


def _format_dup_report(
    csv_path: Path,
    keys_spec: str,
    key_idx: list[int],
    header_mode: bool,
    header_cells: list[str] | None,
    dup_groups: list[tuple[tuple[str, ...], list[int]]],
) -> str:
    lines: list[str] = []
    lines.append(f"CSV: {csv_path.resolve()}")
    lines.append(f"Keys (--keys): {keys_spec.strip()}")
    if header_mode and header_cells is not None:
        lines.append(f"Key column names (from row 1): {', '.join(header_cells)}")
        lines.append(f"Key column indexes (1-based): {', '.join(str(i + 1) for i in key_idx)}")
    else:
        lines.append(f"Key column indexes (1-based): {', '.join(str(i + 1) for i in key_idx)}")
        lines.append("(numeric key mode: every row is data; row 1 is not treated as a header)")
    n_dup_keys = len(dup_groups)
    n_dup_rows = sum(len(rows) for _, rows in dup_groups)
    lines.append(f"Duplicate key values: {n_dup_keys}")
    lines.append(f"Rows participating in duplicate groups: {n_dup_rows}")
    lines.append("")
    if not dup_groups:
        lines.append("No duplicate keys found.")
        return "\n".join(lines)
    for key, row_nums in sorted(dup_groups, key=lambda x: (x[0], x[1][0])):
        lines.append(f"--- Key {key!r}  (count={len(row_nums)}) ---")
        lines.append(f"Row number(s) (1-based): {', '.join(str(n) for n in row_nums)}")
    return "\n".join(lines)


def cmd_dup(csv_path: Path, keys_spec: str, encoding: str, output: Path | None) -> int:
    if not keys_spec or not str(keys_spec).strip():
        print("Error: --keys is required with --dup", file=sys.stderr)
        return 1
    if not csv_path.exists():
        print(f"Error: File not found: {csv_path}", file=sys.stderr)
        return 1
    if not csv_path.is_file():
        print(f"Error: Not a file: {csv_path}", file=sys.stderr)
        return 1
    try:
        rows = _read_csv_rows(csv_path, encoding)
    except OSError as e:
        print(f"Error reading CSV: {e}", file=sys.stderr)
        return 1
    except UnicodeDecodeError as e:
        print(f"Error decoding CSV (try --encoding): {e}", file=sys.stderr)
        return 1
    if not rows:
        text = _format_dup_report(
            csv_path.resolve(),
            keys_spec,
            [],
            False,
            None,
            [],
        )
        if output is not None:
            output.write_text(text + "\n", encoding="utf-8")
            print(f"Wrote report to {output}")
        else:
            print(text)
        return 0

    width = _max_row_width(rows)
    header_mode = not _keys_spec_is_numeric(keys_spec)
    header_cells: list[str] | None = None
    try:
        if header_mode:
            key_idx = _parse_name_key_columns(keys_spec, rows[0])
            header_cells = [rows[0][i] if i < len(rows[0]) else "" for i in key_idx]
            data_start = 1
        else:
            key_idx = _parse_numeric_key_columns(keys_spec, max(width, 1))
            data_start = 0
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    groups: dict[tuple[str, ...], list[int]] = defaultdict(list)
    for ri in range(data_start, len(rows)):
        row = rows[ri]
        key = tuple(row[j] if j < len(row) else "" for j in key_idx)
        groups[key].append(ri + 1)

    dup_groups = [(k, v) for k, v in groups.items() if len(v) > 1]
    text = _format_dup_report(csv_path.resolve(), keys_spec, key_idx, header_mode, header_cells, dup_groups)
    if output is not None:
        output.write_text(text + "\n", encoding="utf-8")
        print(f"Wrote report to {output}")
    else:
        print(text)
    return 0


def _format_table(path: Path, stats: Iterable[SheetStat]) -> str:
    stats = list(stats)
    lines: list[str] = []
    lines.append(f"Workbook: {path}")
    lines.append(f"Sheets: {len(stats)}")
    lines.append("")
    if not stats:
        lines.append("(none)")
        return "\n".join(lines)
    if stats[0].column_labels is not None:
        blocks: list[str] = []
        for s in stats:
            blk: list[str] = []
            blk.append(f"=== Sheet {s.index}: {s.name} ===")
            blk.append(f"Rows: {s.rows}    Columns: {s.columns}")
            if s.columns == 0 or not s.column_labels:
                blk.append("Column names: (none)")
            else:
                blk.append("Column names:")
                blk.append("  " + " | ".join(s.column_labels))
            blocks.append("\n".join(blk))
        lines.append("\n\n".join(blocks))
        return "\n".join(lines)
    header = f"{'#':>3}  {'Sheet name':<40}  {'Rows':>8}  {'Cols':>8}"
    lines.append(header)
    lines.append("-" * len(header))
    for s in stats:
        name = s.name if len(s.name) <= 40 else s.name[:37] + "..."
        lines.append(f"{s.index:>3}  {name:<40}  {s.rows:>8}  {s.columns:>8}")
    return "\n".join(lines)


def cmd_xc(workbook: Path, sheet_tokens: list[str] | None, csv_dir: Path | None) -> int:
    if not workbook.exists():
        print(f"Error: File not found: {workbook}", file=sys.stderr)
        return 1
    if not workbook.is_file():
        print(f"Error: Not a file: {workbook}", file=sys.stderr)
        return 1
    out_dir = csv_dir if csv_dir is not None else workbook.parent
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        print(f"Error: Cannot create output directory {out_dir}: {e}", file=sys.stderr)
        return 1
    try:
        paths = export_workbook_csv(workbook.resolve(), sheet_tokens, out_dir.resolve())
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"Error exporting CSV: {e}", file=sys.stderr)
        return 1
    if not paths:
        if sheet_tokens:
            print(
                "Error: No sheets matched --sheets / -sh (use names, 1-based indexes as in --sum, or #N / @N for index).",
                file=sys.stderr,
            )
        else:
            print("Error: No sheets were exported.", file=sys.stderr)
        return 1
    for p in paths:
        print(p)
    return 0


def cmd_sum(workbook: Path, output: Path | None, sheet_tokens: list[str] | None) -> int:
    if not workbook.exists():
        print(f"Error: File not found: {workbook}", file=sys.stderr)
        return 1
    if not workbook.is_file():
        print(f"Error: Not a file: {workbook}", file=sys.stderr)
        return 1
    try:
        stats, bad = summarize_workbook(workbook, sheet_tokens)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"Error reading workbook: {e}", file=sys.stderr)
        return 1
    if bad:
        print(
            f"Warning: unknown or invalid sheet selector(s), skipped: {', '.join(bad)}",
            file=sys.stderr,
        )
    if sheet_tokens and not stats:
        print(
            "Error: No sheets matched --sheets / -sh (use names, 1-based indexes as in --sum, or #N / @N for index).",
            file=sys.stderr,
        )
        return 1

    text = _format_table(workbook.resolve(), stats)
    if output is not None:
        output.write_text(text + "\n", encoding="utf-8")
        print(f"Wrote summary to {output}")
    else:
        print(text)
    return 0


def cmd_merge_csv(input_dir: Path, output: Path | None, encoding: str, globpat: str) -> int:
    """Merge CSV files in a directory into one CSV beside the directory (default name)."""
    d = input_dir.expanduser()
    if not d.exists():
        print(f"Error: Not found: {input_dir}", file=sys.stderr)
        return 1
    if not d.is_dir():
        print(f"Error: Not a directory: {input_dir}", file=sys.stderr)
        return 1
    d = d.resolve()
    candidates = sorted(d.glob(globpat))
    csv_paths = [p for p in candidates if p.is_file() and p.suffix.lower() == ".csv"]
    out_path = (output if output is not None else d.parent / f"{d.name}_merged.csv").resolve()
    csv_paths = [p for p in csv_paths if p.resolve() != out_path]
    if len(csv_paths) < 1:
        print(f"Error: No .csv files matching {globpat!r} under {d}", file=sys.stderr)
        return 1

    header_ref: list[str] | None = None
    data_rows_written = 0
    first_name = csv_paths[0].name
    # (path_str, rows_in_source_file, data_rows_in_source_excluding_header, rows_written_to_output)
    per_file: list[tuple[str, int, int, int]] = []
    try:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with out_path.open("w", newline="", encoding="utf-8-sig") as outf:
            writer = csv.writer(outf, lineterminator="\n")
            for fi, path in enumerate(csv_paths):
                try:
                    inf = path.open(newline="", encoding=encoding)
                except OSError as e:
                    if fi == 0:
                        print(f"Error: could not open first CSV {path}: {e}", file=sys.stderr)
                        return 1
                    print(f"Warning: could not open {path}: {e}", file=sys.stderr)
                    continue
                try:
                    reader = csv.reader(inf)
                    first = next(reader, None)
                    if first is None:
                        if fi == 0:
                            print(f"Error: First CSV is empty: {path}", file=sys.stderr)
                            return 1
                        continue
                    if fi == 0:
                        header_ref = first
                        writer.writerow(header_ref)
                        data_rows_written += 1
                        n_data_in_file = 0
                        for row in reader:
                            writer.writerow(row)
                            data_rows_written += 1
                            n_data_in_file += 1
                        rows_in_file = 1 + n_data_in_file
                        per_file.append(
                            (str(path.resolve()), rows_in_file, n_data_in_file, rows_in_file)
                        )
                    else:
                        if header_ref is not None and first != header_ref:
                            print(
                                f"Warning: header in {path.name} differs from {first_name}; "
                                "skipping that header line.",
                                file=sys.stderr,
                            )
                        n_data_in_file = 0
                        for row in reader:
                            writer.writerow(row)
                            data_rows_written += 1
                            n_data_in_file += 1
                        rows_in_file = 1 + n_data_in_file
                        per_file.append(
                            (str(path.resolve()), rows_in_file, n_data_in_file, n_data_in_file)
                        )
                finally:
                    inf.close()
    except OSError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"Error merging CSVs: {e}", file=sys.stderr)
        return 1

    n_data_merged = data_rows_written - 1 if data_rows_written > 0 else 0
    print(f"Merged into: {out_path}")
    print(f"Sources merged ({len(per_file)} file(s), in order):")
    for pstr, rows_in_file, data_excl_hdr, written in per_file:
        if written == rows_in_file:
            note = "wrote header + all data rows"
        else:
            note = "skipped duplicate header row; wrote data rows only"
        print(f"  - {pstr}")
        print(f"      Rows in this file: {rows_in_file} (1 header + {data_excl_hdr} data)")
        print(f"      Rows appended to merged output: {written} ({note})")
    print(f"Total data rows in merged file (excluding the single header line): {n_data_merged}")
    print("Output encoding: UTF-8 with BOM.")
    return 0


def main() -> None:
    argv = _normalize_argv(sys.argv[1:])
    parser = argparse.ArgumentParser(
        description="Excel workbook and CSV utilities.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python xlstool.py -sum data.xls
  python xlstool.py --sum report.xlsx
  python xlstool.py --sum book.xlsx -o summary.txt
  python xlstool.py -sum book.xlsx -sh "1,Summary"
  python xlstool.py --sum export.csv
  python xlstool.py -sum data.csv -sh 1
  python xlstool.py -xc book.xlsx
  python xlstool.py --xc data.xls --sheets "Sheet1,Totals"
  python xlstool.py -xc book.xlsx -sh "1,3"
  python xlstool.py --xc book.xlsx --sheets "#2"
  python xlstool.py --xc report.xlsx --csv-dir ./out_csv
  python xlstool.py -dup data.csv --keys "1,3-5"
  python xlstool.py --dup -i data.csv --keys "Name,Email"
  python xlstool.py --dup report.csv --keys "1,2" -o dup_report.txt
  python xlstool.py --merge-csv ./out/bank_parts
  python xlstool.py --merge-csv ./out/bank_parts -o combined.csv
  python xlstool.py --merge-csv ./data --merge-pattern "part-*.csv"
        """,
    )
    parser.add_argument(
        "--sum",
        action="store_true",
        help="List sheet row/column counts for workbooks (tight non-empty box). For .csv, one sheet summary always includes column names from the top row of that box. With --sheets / -sh: filter sheets (workbooks) or the single CSV sheet by stem / 1 / #1.",
    )
    parser.add_argument(
        "--xc",
        action="store_true",
        help="Export sheet(s) to CSV (minimal non-empty rectangle per sheet) as {workbook_stem}_{sheetName}.csv (or --csv-dir)",
    )
    parser.add_argument(
        "--dup",
        action="store_true",
        help="Find duplicate rows in a CSV (requires --keys; use -i or positional FILE for the .csv path)",
    )
    parser.add_argument(
        "--keys",
        metavar="SPEC",
        default=None,
        help="With --dup: key columns as 1-based indexes/ranges (1,3,5-7) or comma-separated header names (row 1 = header)",
    )
    parser.add_argument(
        "--encoding",
        metavar="ENC",
        default=None,
        help="With --dup or --merge-csv: CSV encoding when reading inputs (default utf-8-sig); merge output is always utf-8-sig",
    )
    parser.add_argument(
        "-i",
        "--input",
        metavar="FILE",
        help="Input workbook (.xls, .xlsx, .xlsm), CSV with --sum or --dup",
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        metavar="FILE",
        help="Input workbook or CSV (with --sum / --dup) if not using -i",
    )
    parser.add_argument(
        "-o",
        "--output",
        metavar="FILE",
        help="With --sum, --dup, or --merge-csv: output path (merge: default is <parent>/<dirname>_merged.csv)",
    )
    parser.add_argument(
        "--sheets",
        metavar="NAMES",
        help="With --sum: limit sheets and show column names (top row of tight box). With --xc: export those sheets. #N/@N = index. Same as -sh.",
    )
    parser.add_argument(
        "--csv-dir",
        metavar="DIR",
        help="With --xc: directory for CSV files (default: same directory as the workbook)",
    )
    parser.add_argument(
        "--merge-csv",
        dest="merge_input_dir",
        metavar="DIR",
        default=None,
        help="Merge CSVs in DIR (sorted) into <parent>/<DIRname>_merged.csv; first row is header, rest appended from each file",
    )
    parser.add_argument(
        "--merge-pattern",
        default="*.csv",
        metavar="GLOB",
        help="With --merge-csv: glob under DIR (default: *.csv)",
    )
    args = parser.parse_args(argv)

    mode_n = (
        int(bool(args.sum))
        + int(bool(args.xc))
        + int(bool(args.dup))
        + int(args.merge_input_dir is not None)
    )
    if mode_n != 1:
        parser.error("Specify exactly one of --sum, --xc, --dup, or --merge-csv.")
    if args.sheets and not (args.sum or args.xc):
        parser.error("--sheets / -sh requires --sum or --xc.")
    if args.csv_dir and not args.xc:
        parser.error("--csv-dir is only valid with --xc.")
    if args.encoding is not None and not (args.dup or args.merge_input_dir is not None):
        parser.error("--encoding is only valid with --dup or --merge-csv.")
    if args.keys is not None and not args.dup:
        parser.error("--keys is only valid with --dup.")
    if args.dup and not args.keys:
        parser.error("--dup requires --keys (1-based column indexes/ranges or header names).")
    if args.output and args.xc:
        parser.error("-o/--output is not used with --xc.")
    if args.merge_input_dir is not None:
        if args.sheets or args.keys or args.csv_dir:
            parser.error("--merge-csv cannot be combined with --sheets, --keys, or --csv-dir.")
        if args.input or args.workbook:
            parser.error("With --merge-csv, do not pass -i/--input or a positional FILE.")

    if args.merge_input_dir is not None:
        enc = args.encoding if args.encoding else "utf-8-sig"
        out_path = Path(args.output) if args.output else None
        sys.exit(cmd_merge_csv(Path(args.merge_input_dir), out_path, enc, args.merge_pattern))

    wb = args.input or args.workbook
    if not wb:
        parser.error("Input path required (positional FILE or -i/--input)")

    if args.xc:
        csv_dir = Path(args.csv_dir) if args.csv_dir else None
        sys.exit(cmd_xc(Path(wb), _parse_sheet_tokens(args.sheets), csv_dir))
    out_path = Path(args.output) if args.output else None
    if args.dup:
        enc = args.encoding if args.encoding else "utf-8-sig"
        sys.exit(cmd_dup(Path(wb), args.keys, enc, out_path))
    sys.exit(cmd_sum(Path(wb), out_path, _parse_sheet_tokens(args.sheets)))


if __name__ == "__main__":
    main()
